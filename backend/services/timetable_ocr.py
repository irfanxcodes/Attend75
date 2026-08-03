"""
Timetable OCR & XLSX parser.

Handles three additional input formats beyond PDF:
  - Images (JPG, PNG, WEBP, BMP, TIFF) → OCR via pytesseract + Pillow preprocessing
  - XLSX / XLS spreadsheets → openpyxl table extraction

The output is always a raw text string in the same pipe-delimited or plain-text
format that _parse_timetable_from_text() and the format-auto-detecting
_parse_timetable_pdf() already understand.  When we have a structured table
(XLSX) we also try the direct table parsers (_parse_law_timetable,
_parse_btech_timetable) for cleaner results.
"""

from __future__ import annotations

import io
import logging
import re

logger = logging.getLogger(__name__)

# ── Dependency checks at import time ─────────────────────────────────────────
try:
    import openpyxl as _openpyxl_module   # noqa: F401
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    logging.getLogger(__name__).warning(
        "openpyxl is not installed — XLSX timetable uploads will be unavailable. "
        "Fix: pip install openpyxl==3.1.5"
    )

try:
    from PIL import Image as _PIL_Image   # noqa: F401
    _PILLOW_AVAILABLE = True
except ImportError:
    _PILLOW_AVAILABLE = False

# ── Supported MIME types / extensions ────────────────────────────────────────
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff", ".tif"}
XLSX_EXTENSIONS  = {".xlsx", ".xls"}
ALL_SUPPORTED    = IMAGE_EXTENSIONS | XLSX_EXTENSIONS | {".pdf"}


def is_image(filename: str) -> bool:
    return any(filename.lower().endswith(ext) for ext in IMAGE_EXTENSIONS)


def is_xlsx(filename: str) -> bool:
    return any(filename.lower().endswith(ext) for ext in XLSX_EXTENSIONS)


# ── Image → text (OCR) ───────────────────────────────────────────────────────

def ocr_image_bytes(data: bytes, filename: str = "") -> str:
    """
    Run OCR on an image and return extracted text.

    Uses pytesseract (wraps system Tesseract).  If Tesseract is not installed,
    raises RuntimeError with an install hint so the caller can return a
    user-friendly message.

    Preprocessing pipeline (all via Pillow — no OpenCV required):
      1. Convert to greyscale
      2. Upscale to at least 1800px wide (improves accuracy on phone photos)
      3. Increase contrast
      4. Binarise (threshold → pure black/white)
    """
    try:
        import pytesseract
        from PIL import Image, ImageEnhance, ImageFilter, ImageOps
    except ImportError as exc:
        raise RuntimeError(
            "pytesseract or Pillow not installed. "
            "Run: pip install pytesseract pillow"
        ) from exc

    # Verify tesseract binary is reachable
    try:
        pytesseract.get_tesseract_version()
    except pytesseract.TesseractNotFoundError:
        raise RuntimeError(
            "Tesseract OCR is not installed on this server. "
            "Install it with: sudo apt-get install -y tesseract-ocr"
        )

    from PIL import Image
    img = Image.open(io.BytesIO(data))

    # Convert to RGB first (handles RGBA / palette images)
    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")

    # 1. Greyscale
    grey = img.convert("L")

    # 2. Upscale if too small (min 1800px wide for clean OCR)
    w, h = grey.size
    if w < 1800:
        scale = 1800 / w
        grey = grey.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

    # 3. Enhance contrast
    grey = ImageEnhance.Contrast(grey).enhance(2.0)

    # 4. Sharpen slightly
    grey = grey.filter(ImageFilter.SHARPEN)

    # 5. Binarise with Otsu-like fixed threshold (128 works well for documents)
    grey = grey.point(lambda p: 255 if p > 128 else 0, "1")

    # OCR — use page segmentation mode 6 (single block of text) which works
    # well for timetable grids; lang=eng
    text = pytesseract.image_to_string(
        grey,
        config="--psm 6 --oem 3 -c tessedit_char_whitelist="
               "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz"
               "0123456789 .:/-()&",
    )
    logger.info("OCR completed for %s: %d chars extracted", filename, len(text))
    return text


# ── XLSX → tables ────────────────────────────────────────────────────────────

def parse_xlsx_bytes(data: bytes, filename: str = "") -> tuple[list[list[list]], str]:
    """
    Parse an XLSX/XLS file and return (tables, full_text).

    tables: list of sheets, each sheet is a list of rows, each row a list of cell strings.
    full_text: all cell values joined as plain text (for format detection).
    """
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError(
            "openpyxl is not installed on this server. "
            "Run on the server: pip install openpyxl==3.1.5"
        )

    import openpyxl

    # openpyxl only handles .xlsx. For .xls use xlrd as fallback.
    fname_lower = filename.lower()
    if fname_lower.endswith(".xls") and not fname_lower.endswith(".xlsx"):
        return _parse_xls_bytes(data, filename)

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    all_tables: list[list[list]] = []
    text_parts: list[str] = []

    for ws in wb.worksheets:
        sheet_rows: list[list] = []
        for row in ws.iter_rows(values_only=True):
            row_cells = []
            for cell in row:
                val = "" if cell is None else str(cell).strip()
                row_cells.append(val)
                if val:
                    text_parts.append(val)
            # Skip completely empty rows
            if any(row_cells):
                sheet_rows.append(row_cells)
        if sheet_rows:
            all_tables.append(sheet_rows)

    wb.close()
    full_text = "\n".join(text_parts)
    logger.info("XLSX parsed: %d sheets, %d text chars", len(all_tables), len(full_text))
    return all_tables, full_text


def _parse_xls_bytes(data: bytes, filename: str) -> tuple[list[list[list]], str]:
    """Fallback for legacy .xls files using xlrd."""
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "xlrd not installed (needed for .xls files). Run: pip install xlrd"
        ) from exc

    wb = xlrd.open_workbook(file_contents=data)
    all_tables: list[list[list]] = []
    text_parts: list[str] = []

    for ws in wb.sheets():
        sheet_rows: list[list] = []
        for row_idx in range(ws.nrows):
            row_cells = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                val = str(cell.value).strip() if cell.value else ""
                row_cells.append(val)
                if val:
                    text_parts.append(val)
            if any(row_cells):
                sheet_rows.append(row_cells)
        if sheet_rows:
            all_tables.append(sheet_rows)

    return all_tables, "\n".join(text_parts)


# ── Unified entry point ──────────────────────────────────────────────────────

def extract_timetable_from_upload(
    data: bytes,
    filename: str,
    student_section: str | None = None,
) -> list[dict]:
    """
    Parse any supported file type into a flat list of timetable class entries.

    Dispatch:
      .pdf  → handled by caller (timetable_service._parse_timetable_pdf)
      image → OCR → text → _parse_timetable_from_text / format detectors
      xlsx  → openpyxl → tables → format detectors

    Returns list of dicts with keys: day, time, time_sort, course, section,
    faculty, semester, room, course_key.
    Raises RuntimeError with user-friendly message on unrecoverable errors.
    """
    from services.timetable_service import (
        _parse_timetable_from_text,
        _parse_btech_timetable,
        _parse_law_timetable,
    )

    fname_lower = filename.lower()

    # ── Image path ────────────────────────────────────────────────────────────
    if is_image(fname_lower):
        raw_text = ocr_image_bytes(data, filename)
        if not raw_text.strip():
            raise RuntimeError(
                "OCR couldn't extract any text from this image. "
                "Make sure the photo is clear, well-lit, and not blurry."
            )

        # Try structured parsers first using the OCR text
        classes = _try_all_parsers([], raw_text, student_section)
        if classes:
            return classes

        raise RuntimeError(
            "OCR read the image but couldn't find a timetable structure. "
            "Try uploading a clearer photo or the original PDF/Excel file."
        )

    # ── XLSX path ─────────────────────────────────────────────────────────────
    if is_xlsx(fname_lower):
        all_tables, full_text = parse_xlsx_bytes(data, filename)

        if not full_text.strip():
            raise RuntimeError("The spreadsheet appears to be empty.")

        # Each sheet is treated as a table list
        flat_tables = all_tables  # list of sheets (each is list-of-rows)

        classes = _try_all_parsers(flat_tables, full_text, student_section)
        if classes:
            return classes

        raise RuntimeError(
            "Couldn't find a timetable structure in this spreadsheet. "
            "Make sure it's your class timetable and not a different file."
        )

    # Unsupported type — should not reach here if caller validates first
    raise RuntimeError(f"Unsupported file type: {filename}")


def _try_all_parsers(
    tables: list,
    text: str,
    student_section: str | None,
) -> list[dict]:
    """Try all known timetable format parsers and return the best result."""
    from services.timetable_service import (
        _parse_timetable_from_text,
        _parse_btech_timetable,
        _parse_law_timetable,
    )

    text_upper = text.upper()

    # Law / Session format
    has_session = bool(re.search(r'\bSESSION\s*\d\b', text_upper))
    has_paren_fac = bool(re.search(r'\([A-Z][a-z]', text))

    # B.Tech / grid format
    has_slot = bool(re.search(r'\bS\s*[1-9]\b|\bSlot.?\d\b', text[:800], re.IGNORECASE))
    has_section = bool(re.search(r'\bSection\s*:', text, re.IGNORECASE))

    # Day-grid format (rows start with Mon/Tue/Wed) — works for both B.Tech XLSX and plain grids
    has_day_rows = bool(re.search(r'\b(Mon|Tue|Wed|Thu|Fri)\b', text, re.IGNORECASE))

    results: list[tuple[int, list[dict]]] = []

    if has_session and has_paren_fac:
        r = _parse_law_timetable(tables, text, student_section)
        if r:
            results.append((len(r), r))

    if has_slot or has_section or has_day_rows:
        # Try text-based B.Tech parser first
        r = _parse_btech_timetable(tables, text, student_section)
        if r:
            results.append((len(r), r))

        # Also try direct table-grid parser (works better for XLSX / OCR output
        # where spatial structure isn't preserved as text lines)
        if tables:
            r = _parse_grid_tables(tables, text, student_section)
            if r:
                results.append((len(r), r))

    # Always try generic ASCII/plain-text parser
    r = _parse_timetable_from_text(text)
    if r:
        results.append((len(r), r))

    # Table-grid as last resort when nothing else matched
    if not results and tables:
        r = _parse_grid_tables(tables, text, student_section)
        if r:
            results.append((len(r), r))

    if not results:
        return []

    if student_section:
        # When a section is known, prefer the result with the FEWEST entries
        # that still has at least 5 (1 class/day minimum across a week).
        # More entries = grabbed too many sections. Fewer = properly filtered.
        valid = [(n, r) for n, r in results if n >= 5]
        if valid:
            valid.sort(key=lambda x: x[0])  # ascending = tightest filter
            return valid[0][1]

    # No section filter — return the result with the most entries (widest parse)
    results.sort(key=lambda x: x[0], reverse=True)
    return results[0][1]


def _section_matches(pdf_section: str, student_section: str) -> bool:
    """
    Fuzzy section matching between what's in the PDF and what the portal returns.

    Examples that should match:
      PDF="A"  student="ACC"   → True  (ACC contains A, likely means "Section A CSE")
      PDF="A"  student="A"     → True
      PDF="A1" student="A"     → True  (prefix)
      PDF="B"  student="B1"    → True  (prefix)
      PDF="A"  student="CSE-A" → True  (suffix after dash)
      PDF="A"  student="II-A"  → True  (suffix after dash)
    """
    ps = pdf_section.strip().upper()
    ss = student_section.strip().upper()
    if ps == ss:
        return True
    # Prefix match
    if ps.startswith(ss) or ss.startswith(ps):
        return True
    # Student section has a prefix like "CSE-A", "II-A", "B.TECH-B"
    # Extract the trailing letter/number after the last dash or dot
    for sep in ("-", ".", "_", "/"):
        if sep in ss:
            suffix = ss.split(sep)[-1].strip()
            if suffix and (ps == suffix or ps.startswith(suffix) or suffix.startswith(ps)):
                return True
    # Student section like "ACC" might mean "A" section of CSE
    # Try first character match as last resort only if both are very short
    if len(ps) == 1 and len(ss) >= 1 and ss[0] == ps[0]:
        return True
    return False


def _split_grid_into_section_blocks(grid: list) -> list[tuple[str, list]]:
    """
    Split a single grid that contains multiple section blocks into
    (section_id, rows) pairs.

    A new section starts when a row contains "Section: X" pattern.
    Returns list of (section_id, row_list) tuples.
    If no section headers found, returns [("", grid)] — treat as one block.
    """
    blocks: list[tuple[str, list]] = []
    current_section = ""
    current_rows: list = []

    for row in grid:
        row_text = " ".join(str(c) for c in row if c).strip()
        sec_match = re.search(r'Section\s*[:\-]?\s*([A-Z][A-Z0-9\-]*)', row_text, re.IGNORECASE)

        if sec_match:
            # Save previous block if it had content
            if current_rows:
                blocks.append((current_section, current_rows))
            current_section = sec_match.group(1).strip().upper()
            current_rows = [row]  # include the header row
        else:
            current_rows.append(row)

    # Save last block
    if current_rows:
        blocks.append((current_section, current_rows))

    return blocks if blocks else [("", grid)]


def _parse_grid_tables(
    tables: list,
    full_text: str,
    student_section: str | None,
) -> list[dict]:
    """
    Direct table-grid parser for XLSX / OCR output.

    Key fix: splits multi-section grids into per-section blocks first,
    then only parses the block(s) matching the student's section.
    This prevents the "105 classes/week" dump where all sections are shown.
    """
    DAYS_MAP = {
        "MON": "Monday", "TUE": "Tuesday", "WED": "Wednesday",
        "THU": "Thursday", "FRI": "Friday", "SAT": "Saturday",
        "MONDAY": "Monday", "TUESDAY": "Tuesday", "WEDNESDAY": "Wednesday",
        "THURSDAY": "Thursday", "FRIDAY": "Friday", "SATURDAY": "Saturday",
    }
    NON_CLASS = {
        "L", "U", "N", "C", "H", "LUNCH", "BREAK", "TEA", "LIBRARY",
        "ACTIVITY", "PRAYER", "SPORTS", "FREE", "", "-", "—",
        "PS-1", "PS-2", "PS-III", "PS-I", "PS-II", "PS1", "PS2",
    }
    SLOT_TIMES = [
        "9:30 – 10:20 AM",
        "10:20 – 11:10 AM",
        "11:25 AM – 12:15 PM",
        "12:15 – 1:05 PM",
        "1:45 – 2:35 PM",
        "2:35 – 3:25 PM",
        "3:25 – 4:15 PM",
        "4:15 – 5:05 PM",
    ]

    all_classes = []

    # Normalize: collect all grids from tables input
    grids = []
    for t in tables:
        if not t:
            continue
        if isinstance(t[0], list):
            grids.append(t)

    if not grids:
        return []

    for grid in grids:
        if len(grid) < 2:
            continue

        # ── Split into per-section blocks ────────────────────────────────
        section_blocks = _split_grid_into_section_blocks(grid)

        # ── Select only matching section blocks ──────────────────────────
        if student_section:
            matched_blocks = [
                (sid, rows) for sid, rows in section_blocks
                if not sid or _section_matches(sid, student_section)
            ]
            # If nothing matched, fall back to all blocks
            # (better than showing nothing, but log a warning)
            if not matched_blocks:
                logger.warning(
                    "No section blocks matched student_section=%r "
                    "(available: %s) — using all blocks",
                    student_section,
                    [s for s, _ in section_blocks],
                )
                matched_blocks = section_blocks
        else:
            matched_blocks = section_blocks

        for section_id, block_rows in matched_blocks:
            # Use the actual PDF section id, not the student's portal section
            # This keeps course_key consistent with how _match_student_classes works
            effective_section = section_id or student_section or "A"

            # ── Build faculty lookup from this block ─────────────────────
            faculty_lookup: dict[str, str] = {}
            for row in block_rows:
                if not row or len(row) < 3:
                    continue
                first = str(row[0]).strip()
                if first.isdigit():
                    course = str(row[1]).strip().upper() if len(row) > 1 else ""
                    fac = str(row[2]).strip() if len(row) > 2 else ""
                    if course and fac and len(course) <= 12:
                        faculty_lookup[course] = fac

            # ── Detect orientation ───────────────────────────────────────
            first_col_vals = [str(r[0]).strip().upper() for r in block_rows if r]
            days_in_col0 = sum(1 for v in first_col_vals if v in DAYS_MAP)

            if days_in_col0 < 2:
                continue  # no day rows found in this block

            # ── Find slot header row ─────────────────────────────────────
            header_row_idx = None
            for idx, row in enumerate(block_rows):
                row_text = " ".join(str(c).upper() for c in row if c)
                if re.search(r'\bS\s*1\b|\bSlot.?1\b|\bSESSION\s*1\b', row_text):
                    header_row_idx = idx
                    break

            # ── Build slot → time mapping ────────────────────────────────
            slot_col_times: dict[int, tuple[str, int]] = {}
            slot_order = 0

            if header_row_idx is not None:
                hrow = block_rows[header_row_idx]
                for col_idx, cell in enumerate(hrow):
                    cell_str = str(cell).strip().upper()
                    if not cell_str or cell_str in ("DAY", "SECTION", "YEAR", "DEPT"):
                        continue
                    if cell_str in ("LUNCH", "BREAK", "L", "U", "N", "C", "H", "TEA"):
                        continue
                    # Check if it looks like an actual time value (e.g. "9:30-10:20")
                    if re.search(r'\d{1,2}[:.]\d{2}', cell_str):
                        t = _parse_time_header(cell_str)
                    else:
                        t = SLOT_TIMES[slot_order] if slot_order < len(SLOT_TIMES) else f"Period {slot_order + 1}"
                    slot_col_times[col_idx] = (t, slot_order)
                    slot_order += 1
            else:
                # No header — assign times to all cols after col 0
                if block_rows and block_rows[0]:
                    for col_idx in range(1, len(block_rows[0])):
                        t = SLOT_TIMES[slot_order] if slot_order < len(SLOT_TIMES) else f"Period {slot_order + 1}"
                        slot_col_times[col_idx] = (t, slot_order)
                        slot_order += 1

            if not slot_col_times:
                continue

            # ── Parse day rows ────────────────────────────────────────────
            start_row = (header_row_idx + 1) if header_row_idx is not None else 0
            for row in block_rows[start_row:]:
                if not row:
                    continue
                day_key = str(row[0]).strip().upper()
                day_name = DAYS_MAP.get(day_key)
                if not day_name:
                    continue

                for col_idx, (time_str, sort_order) in slot_col_times.items():
                    if col_idx >= len(row):
                        continue
                    cell_raw = str(row[col_idx]).strip() if row[col_idx] is not None else ""
                    if not cell_raw:
                        continue

                    cell_upper = cell_raw.upper().strip()
                    if cell_upper in NON_CLASS:
                        continue
                    if cell_raw.startswith('<') or cell_raw.endswith('->'):
                        continue

                    course_code, faculty = _split_course_faculty(cell_raw)
                    course_code = course_code.strip().upper()
                    if not course_code or course_code in NON_CLASS or len(course_code) < 2:
                        continue

                    if not faculty:
                        faculty = faculty_lookup.get(course_code, "")

                    all_classes.append({
                        "day": day_name,
                        "time": time_str,
                        "time_sort": f"{sort_order:02d}",
                        "course_key": f"{course_code}-{effective_section}",
                        "course": course_code,
                        "section": effective_section,
                        "faculty": faculty,
                        "semester": "",
                        "room": "",
                    })

    logger.info(
        "Grid table parser: extracted %d entries for section=%s",
        len(all_classes), student_section or "all",
    )
    return all_classes


def _parse_time_header(raw: str) -> str:
    """Parse a time range like '9:30-10:20' or '09.30-10.20' into display format."""
    times = re.findall(r'(\d{1,2})[:.:-](\d{2})', raw)
    if len(times) >= 2:
        h1, m1 = int(times[0][0]), int(times[0][1])
        h2, m2 = int(times[1][0]), int(times[1][1])

        def fmt(h: int, m: int) -> str:
            if h == 0:
                return f"12:{m:02d} AM"
            elif 1 <= h <= 7:
                return f"{h}:{m:02d} PM"
            elif h < 12:
                return f"{h}:{m:02d} AM"
            elif h == 12:
                return f"12:{m:02d} PM"
            else:
                return f"{h - 12}:{m:02d} PM"

        return f"{fmt(h1, m1)} – {fmt(h2, m2)}"
    return raw.strip()


def _split_course_faculty(cell: str) -> tuple[str, str]:
    """
    Split a timetable cell into (course_code, faculty_name).

    Handles formats:
      "CPC (Ms.Barkha)"       → ("CPC", "Ms.Barkha")
      "CPC\\n(Ms.Barkha)"     → ("CPC", "Ms.Barkha")
      "ICP"                   → ("ICP", "")
      "DBMS-A1"               → ("DBMS", "")   ← strip section suffix
    """
    cell = cell.strip()
    if not cell:
        return "", ""

    # Newline split (common in XLSX merged cells)
    parts = cell.split("\n")
    if len(parts) >= 2:
        course = parts[0].strip()
        rest = " ".join(parts[1:])
        fac_match = re.search(r'\((.+?)\)', rest)
        faculty = fac_match.group(1).strip() if fac_match else rest.strip("() ")
        return re.sub(r'\([^)]*\)', '', course).strip(), faculty

    # Inline parentheses: "CPC (Ms.Barkha)"
    inline = re.match(r'^([A-Z][A-Z0-9 &/\-]{0,20}?)\s*\((.+?)\)\s*$', cell.strip())
    if inline:
        return inline.group(1).strip(), inline.group(2).strip()

    # Just a course code (possibly with dash-section like "ICP-A1")
    # Strip trailing section suffix for B.Tech style
    code = re.sub(r'[-_][A-Z0-9]{1,4}$', '', cell.strip())
    code = re.sub(r'[^A-Z0-9&/ ]', '', code.upper()).strip()
    return code, ""
